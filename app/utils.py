import openpyxl
from io import BytesIO
from .models import PropiedadData
import logging
from app.models import PropiedadData, Copropiedad
from app import db
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime
import calendar

def generate_filtered_excel_report(copropiedad_id=None, year=None, month=None):
    
    """Genera un archivo Excel con datos filtrados por copropiedad, año y mes
       Solo accesible para administradores"""
    
# Crear un nuevo libro de trabajo de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Financiero"

    # Construir la consulta base
    query = PropiedadData.query
    
    # Aplicar filtros
    if copropiedad_id:
        query = query.filter_by(copropiedad_id=copropiedad_id)
    
    # Filtrar por año y mes si se proporcionan
    if year and month:
        # Obtener el primer y último día del mes
        last_day = calendar.monthrange(year, month)[1]
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month, last_day)
        
        # Filtrar por fecha de inicio de facturación
        query = query.filter(PropiedadData.fecha_inicio_facturacion >= start_date,
                            PropiedadData.fecha_inicio_facturacion <= end_date)
    
    # Obtener los resultados
    propiedades = query.all()
    
    # Encabezados
    headers = [
        "Copropiedad", "Inmueble", "Matrícula", "Propietario", 
        "Fecha Ingreso", "Fecha Inicio Facturación", "Periodo de Gracia",
        "Valor Presupuesto", "Valor a Pagar Inmueble", 
        "Valor a Pagar Constructora", "Valor a Pagar Propietario"
    ]
    sheet.append(headers)

    # Datos
    for prop in propiedades:
        copropiedad_nombre = prop.copropiedad.nombre if prop.copropiedad else "No asignada"
        propietario = prop.razon_social if prop.tipo_persona == 'Jurídica' else f"{prop.primer_nombre or ''} {prop.primer_apellido or ''}".strip()
        
        row = [
            copropiedad_nombre,
            prop.inmueble,
            prop.matricula,
            propietario,
            prop.fecha_ingreso.strftime('%Y-%m-%d') if prop.fecha_ingreso else "N/A",
            prop.fecha_inicio_facturacion.strftime('%Y-%m-%d') if prop.fecha_inicio_facturacion else "N/A",
            prop.periodo_de_gracia,
            prop.valor_presupuesto,
            prop.valor_a_pagar_inmueble,
            prop.valor_a_pagar_constructora,
            prop.valor_a_pagar_propietario
        ]
        sheet.append(row)

    # Ajustar el ancho de las columnas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar el libro de trabajo en un stream de bytes
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    return excel_stream

def generate_pdf_report(copropiedad_id):
    """Genera un reporte PDF con los datos de los inmuebles de una copropiedad."""
         
    # Obtener la copropiedad
    copropiedad = Copropiedad.query.get(copropiedad_id)
    if not copropiedad:
        raise ValueError("Copropiedad no encontrada")
    
    # Obtener los inmuebles asociados a la copropiedad
    propiedades = PropiedadData.query.filter_by(copropiedad_id=copropiedad_id).all()
    
    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Título
    elements.append(Paragraph(f"Reporte de Inmuebles - {copropiedad.nombre}", title_style))
    elements.append(Spacer(1, 12))
    
    # Información de la copropiedad
    elements.append(Paragraph(f"Dirección: {copropiedad.direccion}", normal_style))
    if copropiedad.nit:
        elements.append(Paragraph(f"NIT: {copropiedad.nit}", normal_style))
    elements.append(Paragraph(f"Fecha del reporte: {datetime.now().strftime('%Y-%m-%d')}", normal_style))
    elements.append(Spacer(1, 12))
    
    # Tabla de inmuebles
    if propiedades:
        # Encabezados
        data = [["Inmueble", "Matrícula", "Propietario", "Fecha Ingreso", "Fecha Inicio Facturación"]]
        
        # Datos
        for prop in propiedades:
            propietario = prop.razon_social if prop.tipo_persona == 'Jurídica' else f"{prop.primer_nombre or ''} {prop.primer_apellido or ''}"
            fecha_ingreso = prop.fecha_ingreso.strftime('%Y-%m-%d') if prop.fecha_ingreso else "N/A"
            fecha_facturacion = prop.fecha_inicio_facturacion.strftime('%Y-%m-%d') if prop.fecha_inicio_facturacion else "N/A"
            
            data.append([
                prop.inmueble,
                prop.matricula,
                propietario.strip(),
                fecha_ingreso,
                fecha_facturacion
            ])
        
        # Crear la tabla
        table = Table(data, repeatRows=1)
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("No se encontraron inmuebles con los criterios seleccionados.", normal_style))
    
    # Generar el PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer           

def generate_excel_file():
    """Genera un archivo Excel con los datos de PropiedadData."""
    
    try:
        # Obtener todos los registros de la base de datos
        propiedades = PropiedadData.query.all()
        
        # Crear un nuevo libro de trabajo de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Datos de Propiedades"

        # Encabezados
        headers = [
            "ID", "Copropiedad", "Inmueble", "Modelo", "Principal", "Agrupar Por", "Matrícula", 
            "Teléfono", "Coeficiente", "Tipo Persona", "Primer Nombre", "Segundo Nombre",
            "Primer Apellido", "Segundo Apellido", "Razón Social", "Tipo ID", 
            "Identificación", "DV", "Email", "Dirección", "Fecha Inicio Facturación", 
            "Fecha Ingreso", "Periodo de Gracia", "Valor a Pagar Inmueble", "Valor Presupuesto", 
            "Valor a Pagar Constructora", "Valor a Pagar Propietario"
        ]
        sheet.append(headers)

        # Datos
        for prop in propiedades:

            copropiedad_nombre = prop.copropiedad.nombre if prop.copropiedad else "No asignada"

            row = [
                prop.id, copropiedad_nombre,prop.inmueble, prop.modelo, prop.principal, prop.agrupar_por, prop.matricula,
                prop.telefono, prop.coeficiente, prop.tipo_persona, prop.primer_nombre, prop.segundo_nombre,
                prop.primer_apellido, prop.segundo_apellido, prop.razon_social, prop.tipo_id,
                prop.identificacion, prop.dv, prop.email, prop.direccion, prop.fecha_inicio_facturacion,
                prop.fecha_ingreso, prop.periodo_de_gracia, prop.valor_a_pagar_inmueble, prop.valor_presupuesto,
                prop.valor_a_pagar_constructora, prop.valor_a_pagar_propietario
            ]
            sheet.append(row)

         # Ajustar el ancho de las columnas para mejor visualización
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Guardar el libro de trabajo en un stream de bytes en memoria
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)  # Mover el cursor al inicio del stream

        return excel_stream
        
    except Exception as e:
        # Registrar el error para depuración
        logging.error(f"Error al generar el archivo Excel: {str(e)}")
        # Re-lanzar la excepción para que sea manejada por la ruta
        raise

def process_excel_upload(file_stream, copropiedad_id):
    """Procesa un archivo Excel y carga los datos en la base de datos."""
    
    workbook = openpyxl.load_workbook(file_stream)
    sheet = workbook.active
    
    # Obtener los encabezados (primera fila)
    headers = [cell.value for cell in sheet[1]]
    
    # Contador de registros procesados y errores
    processed = 0
    errors = 0
    error_messages = []
    
    # Procesar cada fila (excepto la primera que son los encabezados)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        try:
            # Crear un diccionario con los datos de la fila
            data = dict(zip(headers, row))
            
            # Verificar que la matrícula no sea nula
            matricula = data.get('Matrícula')
            if not matricula:
                error_messages.append(f"Fila {row_idx}: El campo 'Matrícula' es obligatorio.")
                errors += 1
                continue
            
            # Verificar si la matrícula ya existe
            if PropiedadData.query.filter_by(matricula=matricula).first():
                error_messages.append(f"Fila {row_idx}: La matrícula {matricula} ya existe.")
                errors += 1
                continue
            
            # Verificar si la identificación ya existe
            identificacion = data.get('Identificación')
            if not identificacion:
                error_messages.append(f"Fila {row_idx}: El campo 'Identificación' es obligatorio.")
                errors += 1
                continue
                
            if PropiedadData.query.filter_by(identificacion=str(identificacion)).first():
                error_messages.append(f"Fila {row_idx}: La identificación {identificacion} ya existe.")
                errors += 1
                continue
            
            # Crear un nuevo objeto PropiedadData
            nueva_propiedad = PropiedadData(
                copropiedad_id=copropiedad_id,
                inmueble=data.get('Inmueble', ''),
                modelo=data.get('Modelo', 'individual'),
                principal=bool(data.get('Principal', False)),
                agrupar_por=data.get('Agrupar Por', 'inmueble'),
                matricula=str(matricula),  # Convertir a string para evitar problemas con números
                telefono=str(data.get('Teléfono', '')),
                coeficiente=float(data.get('Coeficiente', 0) or 0),
                tipo_persona=data.get('Tipo Persona', 'Natural'),
                primer_nombre=data.get('Primer Nombre'),
                segundo_nombre=data.get('Segundo Nombre'),
                primer_apellido=data.get('Primer Apellido'),
                segundo_apellido=data.get('Segundo Apellido'),
                razon_social=data.get('Razón Social'),
                tipo_id=data.get('Tipo ID', 'CC'),
                identificacion=str(identificacion),  # Convertir a string para evitar problemas con números
                dv=str(data.get('DV', '')),
                email=data.get('Email', ''),
                direccion=data.get('Dirección', ''),
                fecha_inicio_facturacion=datetime.strptime(data.get('Fecha Inicio Facturación'), '%Y-%m-%d').date() if data.get('Fecha Inicio Facturación') else None,
                fecha_ingreso=datetime.strptime(data.get('Fecha Ingreso'), '%Y-%m-%d').date() if data.get('Fecha Ingreso') else None,
                periodo_de_gracia=int(data.get('Periodo de Gracia', 0) or 0),
                valor_a_pagar_inmueble=float(data.get('Valor a Pagar Inmueble', 0) or 0),
                valor_presupuesto=float(data.get('Valor Presupuesto', 0) or 0),
                valor_a_pagar_constructora=float(data.get('Valor a Pagar Constructora', 0) or 0),
                valor_a_pagar_propietario=float(data.get('Valor a Pagar Propietario', 0) or 0)
            )
            
            db.session.add(nueva_propiedad)
            processed += 1
            
            # Hacer commit después de cada registro para evitar problemas con transacciones largas
            db.session.commit()
            
        except Exception as e:
            db.session.rollback()
            errors += 1
            error_messages.append(f"Fila {row_idx}: {str(e)}")
    
    return {
        'processed': processed,
        'errors': errors,
        'error_messages': error_messages
    }